import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from xlsxwriter.utility import xl_col_to_name
import xlsxwriter


def read_file_pivot_data(file_name:str):
    """
        This function reads an excel file and creates a dataframe which has the first tow columns as index and the
        rest are pivoted
        :params:file_name: relative path of file
    """
    df = pd.read_excel(file_name)
    df_pivot = df.pivot(index=['Product No', 'Supplier' ], 
                        columns=['Month'], values=['Amount']
                       ).reset_index(
                       ).fillna(0
                       ).droplevel(level=0, axis=1)
    month_column_name = list(df_pivot.columns[2:])
    group_list = ['Product No', 'Supplier' ] + month_column_name
    df_pivot.columns = group_list
    return df_pivot
	
	
def add_rows_for_user_input(df:pd.DataFrame, no_user_input_rows:int):
    """
        This function takes a dataframe and adds the specified number of usr input rows
        :params:df Dataframe to add empty user input rows into
        :params:no_user_input_rows : Number of empty user input rows
    """

    #Add a couple of empty rows
    original_length = df.shape[0]
    num_rows_to_add = 3
    for i in range(num_rows_to_add):
        df.loc[df.shape[0]] = [None] * df.shape[1]

    df.loc[original_length, 'Product No'] = 'User Input'
    return df	


def index_to_excel_col(idx):
    """
        This function converts a number to its equivalent excel column name
    """
    if idx < 1:
        raise ValueError("Index is too small")
    result = ""
    while True:
        if idx > 26:
            idx, r = divmod(idx - 1, 26)
            result = chr(r + ord('A')) + result
        else:
            return chr(idx + ord('A') - 1) + result
			
def complete_df_with_total_rows(df:pd.DataFrame):
    """
        This function creates a total row that sums up all the rows above it (including new user input row values)
        It creates an excel formula and embeds it into the dataframe
        
        :params:df: Dataframe that requires a total row to be created
    """
    df.loc[df.shape[0] - 1, 'Product No'] = "Total"
    num_months = len(df.columns [2:])
    col_name_list = [ index_to_excel_col(i+3) for i in range(num_months)]
    start_row = 2 
    end_row = df.shape[0] - 1
    for col_name, month in zip(col_name_list, list(df.columns [2:])):
        df.loc[df.shape[0] - 1, month] = f"=SUM({col_name}{start_row}:{col_name}{end_row})"
    return df


def write_df(df:pd.DataFrame, file_name:str):
    """
        This function writes the input dataframe to excel workbook whose name is specified
        :params:df Dataframe to be written out
        :file_name:Name of output file to be created
        
    """
    writer = pd.ExcelWriter(file_name, engine='xlsxwriter')

    # Convert the dataframe to an XlsxWriter Excel object.
    df.to_excel(writer, sheet_name='Sheet1', index=False)

    # Close the Pandas Excel writer and output the Excel file.
    writer.save()
    writer.handles = None
    return


def generate_pivot_input_file(input_file_name:str, output_file_name:str):
    df = read_file_pivot_data(input_file_name)
    df = add_rows_for_user_input(df, no_user_input_rows=3)
    df = complete_df_with_total_rows(df)
    write_df(df, output_file_name)
    return


if __name__ == "main":
    ## Call the functions
    generate_pivot_input_file("DynamicPivot-Four Suppliers.xlsx", "Four_Suppliers_Output.xlsx")
    generate_pivot_input_file("DynamicPivot-Five Suppliers.xlsx", "Five_Suppliers_Output.xlsx")
